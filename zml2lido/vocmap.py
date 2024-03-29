"""

Tool that rewrites vocmap data from Excel to xml and vice versa

vocmap -i in.xslx -o out.xml
vocmap -i in.xml -o out.xslx

"""
from lxml import etree
from openpyxl import Workbook
from openpyxl.styles import Font
from pathlib import Path


class Vocmap:
    def __init__(self, *, Input, output):
        Input = Path(Input)
        output = Path(output)
        if Input.suffix == ".xml":
            self.xml2xls(Input=Input, output=output)
        elif Input.suffix == ".xlsx":
            self.xls2xml(Input=Input, output=output)
        else:
            raise SyntaxError("ERROR: Unrecognized input suffix!")

    def xls2xml(self, *, Input, output):
        pass

    def xml2xls(self, *, Input, output):
        """
        Making a new excel from an vocmap.xml file
        """
        doc = etree.parse(str(Input))
        vocL = doc.xpath("/vocmap/voc")
        wb = Workbook()  # make a new Workbook

        # wb.sheetnames to get a list of the sheets
        for voc in vocL:
            name = voc.get("name")
            print(f"voc {name}")
            ws = wb.create_sheet(title=name)
            ws["A1"] = "ID"
            ws["B1"] = "source lang"
            ws["C1"] = "source"
            ws["D1"] = "target"
            ws["E1"] = "scope"
            ws["F1"] = "target lang"
            ws["G1"] = "relation"

            for cell in "A1", "B1", "C1", "D1", "E1", "F1", "G1":
                c = ws[cell]
                c.font = Font(bold=True)

            lno = 1  # line number

            conceptL = voc.xpath("concept")
            for concept in conceptL:
                lno += 1
                print(f" c {lno-1}")
                srcL = concept.xpath("source")
                for src in srcL:
                    slang = src.get("lang")
                    print(f"  s|{slang}:{src.text}")
                    ws[f"A{lno}"] = lno - 1  # concept id
                    ws[f"B{lno}"] = slang
                    ws[f"C{lno}"] = src.text

                trgL = concept.xpath("target")
                for target in trgL:
                    tlang = target.get("lang")
                    scope = target.get("name")
                    print(f"  t|{tlang}:{target.text} from {scope}")
                    ws[f"D{lno}"] = target.text
                    ws[f"E{lno}"] = scope
                    ws[f"F{lno}"] = tlang
                    # ws[f"G{lno}"] = relation

        print(f"... saving to {output}")
        del wb["Sheet"]
        wb.save(filename=output)

    #
    # private helpers
    #
    def _prepare_wb(self, *, output):
        self.wb = Workbook()  # make a new Workbook
        # if output.exists():
        # print(f"Warning: File {output} exists already, will be overwritten!")

        ws1 = self.wb.active
        ws1.title = "MDVOS Liste"
        ws1["A1"] = "objId"
        ws1["B1"] = "IdentNr"
        ws1["C1"] = "Sachbegriff"
        ws1["D1"] = "Titel"
        ws1["E1"] = "Ausstellung [Sektion]"
        ws1.column_dimensions["C"].width = 30
        ws1.column_dimensions["D"].width = 45
        ws1.column_dimensions["E"].width = 90
